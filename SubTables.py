# SubTables.py

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def process_csv_file(file_path, output_directory):
    """
    Process a .csv data file, organize data into separate tables for each assay,
    and save the processed Excel file in the output directory. Recovery values outside 80-120 are highlighted.
    """
    filename = os.path.basename(file_path)

    try:
        # Read .csv file using pandas
        df = pd.read_csv(file_path, skiprows=[0])  # Skip the first row because of Plate Name
        df['% Recovery'] = pd.to_numeric(df['% Recovery'], errors='coerce')

        # Initialize an Excel writer object for the organized data file
        excel_file = os.path.join(output_directory, os.path.splitext(filename)[0] + 'MSDProinflam_organized_data.xlsx')

        # Check if the file already exists
        if os.path.exists(excel_file):
            existing_df = pd.read_excel(excel_file, sheet_name=None)
            existing_data = pd.concat(existing_df.values())
            new_data = pd.concat([df])
            if existing_data.equals(new_data):
                print(f"No changes detected in '{filename}'. Skipping reprocessing.")
                return excel_file, True

        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:  # Use openpyxl engine explicitly
            for assay in df['Assay'].unique():
                assay_df = df[df['Assay'] == assay]

                # Aggregate data for each assay
                table_df = assay_df.groupby('Sample').agg(
                    Assay=('Assay', 'first'),
                    Calc_Concentration_1=('Calc. Concentration', lambda x: x.iloc[0] if len(x) >= 1 else None),
                    Calc_Concentration_2=('Calc. Concentration', lambda x: x.iloc[1] if len(x) >= 2 else None),
                    Avg_Calc_Conc=('Calc. Concentration', 'mean'),
                    Std_Dev_Calc_Conc=('Calc. Concentration', lambda x: x.std() / x.mean() * 100 if len(x) > 1 else None),
                    Recovery_1=('% Recovery', lambda x: x.iloc[0] if len(x) >= 1 else None),
                    Recovery_2=('% Recovery', lambda x: x.iloc[1] if len(x) >= 2 else None),
                    Avg_Recovery=('% Recovery', 'mean'),
                    Std_Dev_Recovery=('% Recovery', lambda x: x.std() / x.mean() * 100 if len(x) > 1 else None),
                ).reset_index()

                # Write to Excel sheet for current assay
                table_df.to_excel(writer, sheet_name=assay, index=False)

                # Apply conditional formatting to % Recovery columns
                ws = writer.sheets[assay]  # Get worksheet by name

                red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')  # Red fill color

                def apply_conditional_formatting(cell):
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        value = float(cell.value)
                        if value < 80 or value > 120:
                            cell.fill = red_fill

                for row in ws.iter_rows(min_row=2, min_col=1, max_col=ws.max_column, max_row=ws.max_row):
                    for cell in row:
                        if cell.column_letter in ['G', 'H']:
                            apply_conditional_formatting(cell)

        print(f"Excel file '{excel_file}' created successfully with assay results organized and formatted.")

        return excel_file, True

    except Exception as e:
        print(f"An error occurred while processing '{filename}': {e}")
        return None, False